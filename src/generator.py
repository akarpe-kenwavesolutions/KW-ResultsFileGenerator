import openpyxl
import os
import copy
from openpyxl.styles import Font, Alignment
from config import Config
from chart_manager import ChartManager


class ResultsGenerator:
    def __init__(self, project_name="KW-Results"):
        self.output_dir = Config.OUTPUT_DIR
        self.template_path = Config.TEMPLATE_PATH
        self.project_name = f"KW-Results-{project_name}"
        self.convert_units = True
        self.wb = None
        self.ws_master = None

    def get_user_preference(self):
        print("-" * 50)
        print("GENERATION SETTINGS")
        print("-" * 50)
        print("Input Data contains both Metric and Imperial values.")
        while True:
            response = input("Select Output Units [1=Imperial (Feet), 2=Metric (Meters)]: ").strip()
            if response == '1':
                self.convert_units = True
                print(">> IMPERIAL (Feet/Inches)")
                break
            elif response == '2':
                self.convert_units = False
                print(">> METRIC (Meters/mm)")
                break

    def load_template(self):
        if not self.template_path:
            raise FileNotFoundError(f"Template file not found in {Config.INPUT_DIR}")
        print(f"\nLoading template: {os.path.basename(self.template_path)}...")
        self.wb = openpyxl.load_workbook(self.template_path)
        self.ws_master = self.wb[self.wb.sheetnames[0]]

    def _sanitize_sheet_name(self, name):
        return name[:31].replace("/", "-").replace("?", "").replace(":", "")

    def _format_ap_name(self, val):
        """Parse access point names as strings, removing decimal points from numeric values."""
        if val is None:
            return None
        try:
            f = float(val)
            if f == int(f):
                return str(int(f))
            return str(f)
        except (ValueError, TypeError):
            return str(val)

    def _convert_thickness(self, val):
        """Handles thickness values (Input assumed mm, convert if Imperial selected)."""
        if val is None:
            return None
        try:
            val = float(val)
            if self.convert_units:
                result = val / 25.4
            else:
                result = val
            return round(result, 3)
        except (ValueError, TypeError):
            return val

    def _format_pipe_spec(self, diameter_str, material_str):
        """
        Format a single pipe spec with proper units.
        Returns formatted string like "315mm PVC" or "12in PVC"
        """
        if not diameter_str and not material_str:
            return ""

        formatted_diameter = ""
        if diameter_str:
            try:
                diameter_mm = float(diameter_str)
                if self.convert_units:
                    diameter_inches = round(diameter_mm / 25.0)
                    formatted_diameter = f"{diameter_inches}in"
                else:
                    formatted_diameter = f"{diameter_str}mm"
            except ValueError:
                formatted_diameter = diameter_str

        result = f"{formatted_diameter} {material_str}".strip()
        return result

    def _format_step_pipe_type(self, spec_tuple):
        """
        Format a (diameter, material) tuple into a display string with correct units.
        Returns None if spec_tuple is None (transition step).
        """
        if spec_tuple is None:
            return None
        dia, mat = spec_tuple
        try:
            dia_float = float(dia)
            if self.convert_units:
                dia_formatted = f"{round(dia_float / 25.0)}in"
            else:
                dia_formatted = f"{int(dia_float)}mm"
        except:
            dia_formatted = dia
        return f"{dia_formatted} {mat}".strip()

    def process_site(self, site_data):
        site_name = self._sanitize_sheet_name(site_data.get('site_name', 'Unknown'))
        print(f"Processing Sheet: {site_name}")
        ws_new = self.wb.copy_worksheet(self.ws_master)
        ws_new.title = site_name

        if hasattr(self.ws_master, '_charts') and self.ws_master._charts:
            try:
                ws_new.add_chart(copy.deepcopy(self.ws_master._charts[0]))
            except:
                pass

        # --- PIPE SPECS: Calculate how many rows to insert ---
        pipe_specs_list = site_data.get('pipe_specs_list', [])
        num_pipe_specs = len(pipe_specs_list)
        num_rows_to_insert = max(0, num_pipe_specs - 1)

        if num_rows_to_insert > 0:
            ws_new.insert_rows(Config.ROW_PIPE_TYPE_LABEL, num_rows_to_insert)

        # --- SITE NAME ---
        ws_new.cell(row=Config.ROW_SITE_NAME, column=Config.COL_SITE_NAME).value = site_name

        # --- PIPE SPECS LIST (bold, ending at Pipe Type label row) ---
        pipe_type_label_row = Config.ROW_PIPE_TYPE_LABEL + num_rows_to_insert
        first_spec_row = pipe_type_label_row - num_pipe_specs + 1

        ws_new.cell(row=pipe_type_label_row, column=Config.COL_PIPE_TYPE_LABEL).value = "Pipe Type"

        current_pipe_row = first_spec_row
        for diameter, material in pipe_specs_list:
            formatted_spec = self._format_pipe_spec(diameter, material)
            if formatted_spec:
                cell = ws_new.cell(row=current_pipe_row, column=Config.COL_PIPE_TYPE)
                cell.value = formatted_spec
                cell.font = Font(bold=True)
                current_pipe_row += 1

        # --- METADATA (adjusted for inserted rows) ---
        row_start_ap = Config.ROW_START_AP_VAL + num_rows_to_insert
        row_end_ap = Config.ROW_END_AP_VAL + num_rows_to_insert
        row_resolution = Config.ROW_RESOLUTION_VAL + num_rows_to_insert

        ws_new.cell(row=row_start_ap, column=Config.COL_START_AP_VAL).value = site_data.get('ap_id_1')
        ws_new.cell(row=row_end_ap, column=Config.COL_END_AP_VAL).value = site_data.get('ap_id_2')
        ws_new.cell(row=row_resolution, column=Config.COL_RESOLUTION_VAL).value = site_data.get('resolution')

        # --- TABLE HEADERS (adjusted for inserted rows) ---
        table_header_row = Config.TABLE_HEADER_ROW + num_rows_to_insert

        if self.convert_units:
            ws_new.cell(row=table_header_row, column=Config.COL_START_FT).value = "Start (ft)"
            ws_new.cell(row=table_header_row, column=Config.COL_END_FT).value = "End (ft)"
        else:
            ws_new.cell(row=table_header_row, column=Config.COL_START_FT).value = "Start (m)"
            ws_new.cell(row=table_header_row, column=Config.COL_END_FT).value = "End (m)"

        # --- PIPE TYPE COLUMN HEADER ---
        # Col H (8) if asset IDs included, Col G (7) if not
        pipe_type_col = Config.COL_ASSET_ID + 1 if Config.REQUIRE_ASSET_IDS else Config.COL_ASSET_ID

        if not Config.REQUIRE_ASSET_IDS:
            ws_new.cell(row=table_header_row, column=Config.COL_ASSET_ID).value = None

        ws_new.cell(row=table_header_row, column=pipe_type_col).value = "Pipe Type"

        # --- DATA TABLE ---
        segments = site_data.get('segments', [])
        data_start_row = Config.DATA_START_ROW + num_rows_to_insert

        # Clear template placeholder data
        for r in range(data_start_row, data_start_row + 500):
            for c in range(1, 9):
                ws_new.cell(row=r, column=c).value = None

        center_align = Alignment(horizontal='center')
        max_midpoint = 0

        current_row = data_start_row
        for seg in segments:

            # --- Column A: Access Point Label (centre aligned) ---
            cell_ap = ws_new.cell(row=current_row, column=Config.COL_ACCESS_POINT_LBL)
            cell_ap.value = seg.get('access_point_label')
            cell_ap.alignment = center_align

            # --- Units & decimal places ---
            if self.convert_units:
                start_val = seg.get('start_ft')
                end_val = seg.get('end_ft')
                decimals = 1
            else:
                start_val = seg.get('start_m')
                end_val = seg.get('end_m')
                decimals = 3

            start_rounded = round(float(start_val), decimals) if start_val is not None else None
            end_rounded = round(float(end_val), decimals) if end_val is not None else None

            # --- Column B: Start (centre aligned) ---
            cell_start = ws_new.cell(row=current_row, column=Config.COL_START_FT)
            cell_start.value = start_rounded
            cell_start.alignment = center_align

            # --- Column C: End (centre aligned) ---
            cell_end = ws_new.cell(row=current_row, column=Config.COL_END_FT)
            cell_end.value = end_rounded
            cell_end.alignment = center_align

            # --- Column D (gray): Midpoint average of start and end ---
            if start_rounded is not None and end_rounded is not None:
                midpoint = round((start_rounded + end_rounded) / 2, decimals)
                ws_new.cell(row=current_row, column=4).value = midpoint
                if midpoint > max_midpoint:
                    max_midpoint = midpoint

            # --- Column E: DRI Thickness ---
            ws_new.cell(row=current_row, column=Config.COL_DRI_THICKNESS).value = self._convert_thickness(
                seg.get('dri_thickness'))

            # --- Column F: Nominal Thickness ---
            ws_new.cell(row=current_row, column=Config.COL_NOM_THICKNESS).value = self._convert_thickness(
                seg.get('nom_thickness'))

            # --- Column G: Asset ID (if required) ---
            if Config.REQUIRE_ASSET_IDS:
                ws_new.cell(row=current_row, column=Config.COL_ASSET_ID).value = seg.get('pipe_asset_id')

            # --- Column G or H: Pipe Type per step ---
            pipe_type_val = self._format_step_pipe_type(seg.get('pipe_type_step'))
            ws_new.cell(row=current_row, column=pipe_type_col).value = pipe_type_val

            current_row += 1

        # --- UPDATE CHART ---
        ordered_segments = site_data.get('ordered_segments', '')
        ChartManager.update_chart_range(
            ws_new,
            len(segments),
            chart_title=ordered_segments,
            data_start_row=data_start_row,
            row_offset=num_rows_to_insert,
            x_axis_max=max_midpoint
        )

    def save(self, input_site_names):
        sanitized_names = [self._sanitize_sheet_name(n) for n in input_site_names]
        if self.ws_master.title not in sanitized_names:
            del self.wb[self.ws_master.title]
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
        full_path = os.path.join(self.output_dir, f"{self.project_name}.xlsx")
        print(f"\nSaving results to: {full_path}")
        self.wb.save(full_path)
        print("Generation Complete.")

    def run(self, data):
        self.get_user_preference()
        self.load_template()
        for site in data:
            self.process_site(site)
        self.save([s['site_name'] for s in data])
